"""Excel (.xlsx) loader — every sheet, one text block per row.

Each row is rendered as ``Sheet: <name>\ncolumn: value\ncolumn: value...``
(column headers repeated per row) instead of a raw tab/comma-separated
table, so a chunk boundary landing mid-table still keeps each value legible
next to the column it belongs to — the same problem CSV ingestion already
solves via LangChain's ``CSVLoader``, applied here per-sheet.
"""

import openpyxl
from langchain_core.documents import Document as LangChainDocument
from langchain_text_splitters import RecursiveCharacterTextSplitter


class ExcelLoader:
    """Loads every sheet of an .xlsx workbook and splits it into chunks."""

    def __init__(self, splitter: RecursiveCharacterTextSplitter) -> None:
        self._splitter = splitter

    def load(self, file_path: str) -> list[LangChainDocument]:
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        documents: list[LangChainDocument] = []
        try:
            for sheet_name in workbook.sheetnames:
                documents.extend(self._load_sheet(workbook[sheet_name], sheet_name))
        finally:
            workbook.close()
        return self._splitter.split_documents(documents)

    @staticmethod
    def _load_sheet(worksheet, sheet_name: str) -> list[LangChainDocument]:
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []

        columns = [str(cell) if cell is not None else "" for cell in header]
        documents: list[LangChainDocument] = []
        for row in rows:
            if all(cell is None for cell in row):
                continue
            lines = [f"Sheet: {sheet_name}"]
            for column, value in zip(columns, row):
                if value is None:
                    continue
                lines.append(f"{column}: {value}")
            if len(lines) == 1:
                continue
            documents.append(
                LangChainDocument(
                    page_content="\n".join(lines), metadata={"sheet": sheet_name}
                )
            )
        return documents
